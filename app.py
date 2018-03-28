# 403 - код ошибки, когда бот не может достучаться до юзера
'''
TODO: переосмыслить смысл таблицы switch:
Мб день недели считать по ходу дела.
Каждый месяц менять фон - фигня. Мб просто хранить текущий.

Убери костыль с fetchone (цикл for). Достаточно проверять кортеж и вытаскивать 1й элемент
'''

import datetime
import logging
import os
import re
import sys
import time
import traceback

from flask import Flask
import openpyxl
import pytz
import telebot
from PIL import Image, ImageDraw, ImageFont

import constants
import functions

bot = telebot.TeleBot(constants.token)
tz = pytz.timezone('Asia/Almaty')

application = Flask(__name__)


@application.route('/')
def hello_world():
    return 'hello world!'


def log(module, info):
    with open('{0}_err.txt'.format(module), 'w') as f:
        f.write(str(datetime.datetime.now(tz)) + '\n' * 2)
        f.writelines(info)
    with open('{0}_err.txt'.format(module), 'rb') as f:
        bot.send_document(constants.main_adm, f)
    os.remove('{0}_err.txt'.format(module))


@bot.message_handler(commands=["ping"])
def pong(msg):
    bot.reply_to(msg, 'pong')


"""
заменить dic на что-то красивое.
Он используется в расписании по inline конпкам
"""
dic = {
    'mon': 'Понедельник',
    'tue': 'Вторник',
    'wed': 'Среду',
    'thu': 'Четверг',
    'fri': 'Пятницу',
    'sat': 'Субботу',
}
days_for_rasp = [
    ('mon', 'Понедельник'),
    ('tue', 'Вторник'),
    ('wed', 'Среду'),
    ('thu', 'Четверг'),
    ('fri', 'Пятницу'),
    ('sat', 'Субботу'),
    ('mon', 'Понедельник'),
]

logger = telebot.logger
logger_level = logging.ERROR
telebot.logger.setLevel(logger_level)

logging.basicConfig(format=u'%(filename)s[LINE:%(lineno)d]# %(levelname)-8s [%(asctime)s]  %(message)s',
                    filename='logs\\{:%Y-%m-%d_%H-%M-%S}.log'.format(datetime.datetime.now(tz=tz)),
                    filemode='w',
                    level=logger_level)


@bot.message_handler(commands=['test'])
def test(message):
    try:
        bot.send_message(message.chat.id, 'Server time ' + str(datetime.datetime.now()))
    except:
        traceback.print_exc()
        ei = "".join(traceback.format_exception(*sys.exc_info()))
        log('test', ei)


@bot.message_handler(commands=['keyboard'])
def kb(message):
    keyboard = telebot.types.ReplyKeyboardMarkup(True)
    keyboard.add(telebot.types.KeyboardButton('\U0001F392 Расписание'), telebot.types.KeyboardButton('\U0001F393 Группа'),
                 telebot.types.KeyboardButton('\U00002753 Помощь'))
    bot.send_message(message.chat.id, 'Чтобы убрать клавиатуру напишите /hide', reply_markup=keyboard)


@bot.message_handler(commands=['hide'])
def hide(message):
    bot.send_message(message.chat.id, 'Клавиатура скрыта', reply_markup=telebot.types.ReplyKeyboardHide())


@bot.message_handler(commands=['video'])
def vid_init(message):
    bot.send_message(message.chat.id, '/vid720 - для получения видео в формате *720p* (Весит _476_ мб)\n/vid360 - для получения видео в *360p* (Весит _116_ мб)',
                     parse_mode='Markdown')


@bot.message_handler(commands=['vid720'])
def vid720(message):
    bot.send_document(message.chat.id, 'BQADAgAD4AADboTaCrqhBUqOVpZ3Ag')


@bot.message_handler(commands=['vid360'])
def vid360(message):
    bot.send_document(message.chat.id, 'BQADAgAD4QADboTaCjS8U8AuMgteAg')


@bot.message_handler(commands=['start'])  # first command to bot / Первая команда боту
def start(message):
    try:
        if functions.isprivate(message):
            try:  # trying to insert information about user. / попытка внести информацию о юзере
                conn = functions.start_sql()
                cursor = conn.cursor(buffered=True)
                insert = "INSERT INTO  `known_users` ( `id` ) VALUES ( {0} )".format(message.from_user.id)
                cursor.execute(insert)
                conn.commit()
                cursor.close()
                conn.close()
            except:
                pass
            keyboard = telebot.types.ReplyKeyboardMarkup(True)
            keyboard.add(telebot.types.KeyboardButton('\U0001F392 Расписание'), telebot.types.KeyboardButton('\U0001F393 Группа'),
                         telebot.types.KeyboardButton('\U00002753 Помощь'))
            bot.send_message(message.chat.id, constants.start.format(message.from_user.first_name), parse_mode='HTML', reply_markup=keyboard)
    except:
        traceback.print_exc()
        ei = "".join(traceback.format_exception(*sys.exc_info()))
        name = message.from_user.first_name + ' ' + message.from_user.last_name + ' ' + message.from_user.username + ' ' + str(message.from_user.id) + '\n'
        log('start ', name + ei)


@bot.message_handler(commands=['stop'])  # deleting from db
def stop(message):
    try:
        if functions.isprivate(message):
            conn = functions.start_sql()
            cursor = conn.cursor()
            try:
                cursor.execute('DELETE FROM `known_users` WHERE id = {0}'.format(message.from_user.id))
                conn.commit()
                bot.send_message(message.from_user.id, 'Вы успешно удалены из базы данных\nЧтобы вернуться, напишите мне /start',
                                 reply_markup=telebot.types.ReplyKeyboardHide())
            except:
                bot.send_message(message.from_user.id, 'Вы не найдены в базе данных')
    except:
        traceback.print_exc()
        ei = "".join(traceback.format_exception(*sys.exc_info()))
        name = message.from_user.first_name + ' ' + message.from_user.last_name + ' ' + message.from_user.username + ' ' + str(message.from_user.id) + '\n'
        log('stop ', name + ei)


@bot.message_handler(func=lambda x: x.text in ['\U0001F393 Группа', '/group'])
# setup user's group / Установка группы юзера
def group(message):
    if functions.isprivate(message):
        try:  # generating and sending keyboard with faculties / Создает клаву с факультетами и отправляет её
            inline_keyboard = telebot.types.InlineKeyboardMarkup(row_width=2)
            buttons = [
                ('Мат-ком', 'М'),
                ('Мат-эконом', 'Э'),
                ('Хим-био', 'Х'),
                ('Физ-мат', 'ФМ')
            ]
            buttons = [telebot.types.InlineKeyboardButton(text=name, callback_data='g_' + char) for name, char in buttons]
            inline_keyboard.add(*buttons)
            bot.send_message(message.chat.id, "С какого вы факультета ?", reply_markup=inline_keyboard)  # from_user
        except:
            traceback.print_exc()
            ei = "".join(traceback.format_exception(*sys.exc_info()))
            name = message.from_user.first_name + ' ' + message.from_user.last_name + ' ' + message.from_user.username + ' ' + str(message.from_user.id) + '\n'
            log('group ', name + ei)


@bot.message_handler(func=lambda x: x.text in ['\U0001F392 Расписание', '/rasp'])
# Sending schedule for today / Отправляет расписание на сегодня
def rasp(message):
    '''
    TODO узнавать день недели по сообщению
    есть поле
    date	Integer	 Date the message was sent in Unix time
    :param message:
    :return:
    '''
    try:
        conn = functions.start_sql()
        cursor = conn.cursor()
        get_klass = 'SELECT klass FROM known_users WHERE id = {0}'.format(message.from_user.id)  # получает группу для клавы
        cursor.execute(get_klass)
        klass = cursor.fetchone()
        if klass:
            klass = klass[0]
            today = (datetime.datetime.now(tz=tz) + datetime.timedelta(hours=8)).weekday()
            cursor.execute("SELECT {0} FROM schedule WHERE klass = '{1}'".format(days_for_rasp[today][0], klass))
            schedule = cursor.fetchone()[0]
            first_text = "Расписание уроков для группы {group} на {day}".format(group=klass, day=days_for_rasp[today][1])
            cursor.execute('SELECT value FROM switch where var="ad" ')
            ad = "\n\n\U0001F50AИнформацию о появлении расписания и не только можно будет найти на канале @school134_info"  # TODO: Сделать возможность менять это сообщение из базы
            schedule_keyboard = functions.schedule_inline_keyboard(days_for_rasp[today][0], klass)
            bot.send_photo(chat_id=message.chat.id, photo=schedule, caption=first_text + ad, reply_markup=schedule_keyboard, disable_notification=True)
        else:
            group(message)
    # except TypeError:
    #    pass
    except:
        bot.reply_to(message, "Напишите мне в лс (@school134_bot) команду /start")
        traceback.print_exc()
        ei = "".join(traceback.format_exception(*sys.exc_info()))
        name = message.from_user.first_name + ' ' + message.from_user.last_name + ' ' + message.from_user.username + ' ' + str(message.from_user.id) + '\n'
        log('rasp ', name + ei)


@bot.message_handler(func=lambda x: x.text in ['\U00002753 Помощь', '/help'])
# just help message
def help(message):
    bot.send_message(message.chat.id, "/start - Начать работу\n"
                                      "/stop - Удалить себя из базы данных\n"
                                      "/group - Изменяет вашу группу\n"
                                      "/rasp - Показывает расписание на _сегодняшний_ день.\n"
                                      "/fb [текст] - Отправляет ваше сообщение разработчику _(используйте её если у вас возникли проблемы или вопросы)_\n"
                                      "/help - Выводит это сообщение\n"
                                      "/video - Скачать фильм-презентацию школы\n", parse_mode='Markdown')


@bot.message_handler(commands=['ahelp'])  # help for admins / такой же хелп, но для админов
def admcmd(message):
    if functions.admincheck(message):
        bot.send_message(message.chat.id, "/ras {exp} - рассылка по сравнению. SELECT id from known_users WHERE {exp}\n"
                                          "/list - Показать список зафиксированных пользователей\n"
                                          "/fblist - Показать список всех фидбэков\n"
                                          "/answer {fb_id} {текст} - Ответить на фидбэк\n"
                                          "/background - Установить фон на следующий месяц (если нет фона - на этот) (нужно делать КАЖДЫЙ месяц!)\n"
                                          "/webhook - Просмотреть информацию о webhook-е\n"
                                          "/sql {query} - Выполнить SQL-запрос в БД")


@bot.message_handler(commands=['addrasp'])  # add addrasp flag in DB / добавляет addrasp флаг с днем недели в бд
def addrap(message):
    if message.from_user.id == constants.main_adm:
        functions.set_next_step('addrasp', message.from_user.id)
        bot.send_message(message.from_user.id, "Можете отправлять файл.")
    else:
        bot.reply_to(message, "Эта функция доступна *ТОЛЬКО* разработчику", parse_mode="Markdown")


@bot.message_handler(commands=['fb'])  # sending feedback to developer / отправляет фидбэк разрабу
def feedback(message):
    try:
        msg = message.text.replace('/fb ', '')
        if msg != message.text:  # if '/fb ' deleted then user sends some text / если сообщение не изменилось, то пользователь просто отправил '/fb' - без текста
            report = "Сообщение от {first_name} {last_name} (@{username})\n" \
                     "Прислано {time}\n" \
                     "Текст: {text}".format(first_name=message.from_user.first_name, last_name=message.from_user.last_name, username=message.from_user.username,
                                            time=str(datetime.datetime.now(tz)),
                                            text=msg)
            conn = functions.start_sql()
            cursor = conn.cursor(buffered=True)
            query = "INSERT INTO feedback (u_id, m_id, text) VALUES (%s, %s, %s)"  # adding feedback to the DB / добавляет фидбэк в бд
            values = (message.from_user.id, message.message_id, report)
            cursor.execute(query, values)
            conn.commit()
            bot.send_message(constants.main_adm, report + '\n/answer ' + str(cursor.lastrowid) + ' текст')  # getting fb_id which auto increments and sending fb to dev
            # Получает id последнего фидбэка
            bot.reply_to(message, "Ваше сообщение отправлено!")
            cursor.close()
            conn.close()
        else:
            bot.reply_to(message, "Команды работает так:\n/fb текст")
    except:
        traceback.print_exc()
        ei = "".join(traceback.format_exception(*sys.exc_info()))
        name = message.from_user.first_name + ' ' + message.from_user.last_name + ' ' + message.from_user.username + ' ' + str(message.from_user.id) + '\n'
        log('fb ', name + ei)


@bot.message_handler(commands=['answer'])  # Ответ на фидбэк
def answer(message):
    try:
        if functions.admincheck(message):
            msg = message.text.replace("/answer ", "")
            if msg != message.text:  # Проверка на случайное нажатие /answer
                conn = functions.start_sql()
                cursor = conn.cursor(buffered=True)
                cursor.execute("SELECT u_id, m_id FROM feedback WHERE id = {0} ".format(msg[0:msg.find(" ")]))  # Вырежет из сообщения номер
                for u_id, m_id in cursor.fetchmany(1):  # Если получать кортеж через fetchone, то вернет итерируемый лист, тогда не будет работать цикл
                    bot.send_message(u_id, "Ответ разработчика:\n" + msg[msg.find(" ") + 1:], reply_to_message_id=m_id)
                    cursor.execute("DELETE FROM feedback WHERE id = {0}".format(msg[0:msg.find(" ")]))  # Удаляет фидбэк, на который был получен ответ, из бд
                    conn.commit()
                cursor.close()
                conn.close()
                bot.reply_to(message, "Сообщение отправлено")
            else:
                bot.reply_to(message, "Команда работает так:\n/answer {fb_id} {text}")
    except:
        traceback.print_exc()
        ei = "".join(traceback.format_exception(*sys.exc_info()))
        name = message.from_user.first_name + ' ' + message.from_user.last_name + ' ' + message.from_user.username + ' ' + str(message.from_user.id) + '\n'
        log('answer ', name + ei)


@bot.message_handler(commands=['fblist'])  # Получает список фидбэков
def feedback_list(message):
    try:
        if functions.admincheck(message):
            conn = functions.start_sql()
            cursor = conn.cursor()
            cursor.execute("SELECT id, text FROM feedback ORDER BY id DESC")  # Сортировка в обратном порядке, чтобы старые фидбэки было удобнее смотреть
            for num, text in functions.iter_row(cursor, 10):
                bot.send_message(message.chat.id, text + '\n/answer ' + str(num) + ' текст')
            bot.send_message(message.chat.id, "Все фидбэки отправлены")
    except:
        traceback.print_exc()
        ei = "".join(traceback.format_exception(*sys.exc_info()))
        name = message.from_user.first_name + ' ' + message.from_user.last_name + ' ' + message.from_user.username + ' ' + str(message.from_user.id) + '\n'
        log('fb list ', name + ei)


@bot.message_handler(commands=['list'])  # Получает список всех пользователей, которые работали с ботом
def user_list(message):
    try:
        if functions.admincheck(message):
            conn = functions.start_sql()
            cursor = conn.cursor(buffered=True)
            cursor.execute("SELECT id, klass FROM known_users ORDER BY klass")
            num = 1
            text = ""
            for id, klass in functions.iter_row(cursor, 50):
                msg = "{0}--{1}\n".format(id, klass)
                text += msg
                if num % 50 == 0:
                    bot.send_message(message.chat.id, text, disable_notification=True)
                    text = ""
                num += 1
            if text != "":
                bot.send_message(message.chat.id, text, disable_notification=True)
            bot.send_message(message.chat.id, "{0} записей в базе".format(num - 1), disable_notification=True)
            cursor.close()
            conn.close()
    except:
        traceback.print_exc()
        ei = "".join(traceback.format_exception(*sys.exc_info()))
        name = message.from_user.first_name + ' ' + message.from_user.last_name + ' ' + message.from_user.username + ' ' + str(message.from_user.id) + '\n'
        log('user list ', name + ei)


@bot.message_handler(commands=['ras'])
def ras_init(message):
    if functions.admincheck(message):
        query = message.text.replace('/ras ', '')
        if query != message.text:
            functions.set_next_step("ras_" + query, message.from_user.id)
            bot.send_message(message.chat.id, "Запрос:\nSELECT id FROM known_users WHERE " + query + "\n\nОтправьте сообщение для рассылки")
        else:
            bot.reply_to(message, "Команды работает так:\n/ras {expression}.\nSELECT id,name FROM known_users WHERE {expression}")


@bot.message_handler(commands=['background'])
def background(message):
    if functions.admincheck(message):
        functions.set_next_step('background-in', message.from_user.id)
        bot.send_message(message.chat.id, "Отправьте фотографию на фон размером ~1280*800")


@bot.message_handler(commands=['webhook'])
def webhook_info(message):
    if functions.admincheck(message):
        try:
            global tz
            info = bot.get_webhook_info()
            msg = "URL: _{0}_\n" \
                  "Custom certificate: _{1}_\n" \
                  "Pending update count: _{2}_\n" \
                  "Last error date: _{3}_\n" \
                  "Last error message: _{4}_".format(info.url, info.has_custom_certificate, info.pending_update_count,
                                                     str(datetime.datetime.fromtimestamp(info.last_error_date, tz)),
                                                     info.last_error_message)
            bot.send_message(message.chat.id, msg, parse_mode="Markdown")
        except:
            traceback.print_exc()
            ei = "".join(traceback.format_exception(*sys.exc_info()))
            name = message.from_user.first_name + ' ' + message.from_user.last_name + ' ' + message.from_user.username + ' ' + str(message.from_user.id) + '\n'
            log('Webhook', name + ei)


@bot.message_handler(commands=['sql'])
def sql(message):
    if functions.admincheck(message):
        try:
            conn = functions.start_sql()
            cursor = conn.cursor()
            cmd = message.text.replace("/sql ", "")  # type: str
            if cmd:
                mul = bool(cmd.count(';'))
                cursor.execute(cmd, multi=mul)
                report = ""
                for elem in functions.iter_row(cursor, 20):
                    report += str(elem) + ' '
                report = "done" if report == '' else report
                bot.reply_to(message, report)
                conn.commit()
                cursor.close()
                conn.close()
            else:
                bot.reply_to(message, "Команда работает так:\n/sql {sql query}")
        except:
            traceback.print_exc()
            ei = "".join(traceback.format_exception(*sys.exc_info()))
            name = message.from_user.first_name + ' ' + message.from_user.last_name + ' ' + message.from_user.username + ' ' + str(message.from_user.id) + '\n'
            log('sql', name + ei)


@bot.message_handler(func=lambda m: True, content_types=['text', 'audio', 'document', 'photo', 'sticker', 'video', 'voice', 'location', 'contact'])
def switch(message):
    try:
        data = functions.select_next_step(message)
        if data != '' and data is not None:
            data_check = data.split('_')
            funcs = {
                'background-in': set_background,
                'addrasp': rasp_add,
                'ras': ras_switch}
            functions.set_next_step('Null', message.from_user.id)
            funcs.get(data_check[0])(message, data)
    except:
        traceback.print_exc()
        ei = "".join(traceback.format_exception(*sys.exc_info()))
        name = message.from_user.first_name + ' ' + message.from_user.last_name + ' ' + message.from_user.username + ' ' + str(message.from_user.id) + '\n'
        log('switch ', name + ei)


@bot.callback_query_handler(func=lambda call: True)
def callback_inline(call):
    if call.message:
        # TODO: Чекни, не мутная ли это схема
        pref = call.data.split('_')  # Разделение в лист по ключу
        # =========================================================================================================== SCHEDULE
        if pref[0] == 's':  # s_{day}_{group}
            if pref[1] == 'rings':  # Звонки берутся с локального файла
                keyboard = functions.schedule_inline_keyboard(pref[1], pref[2])
                if call.message.text:
                    bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id, text=constants.rings, reply_markup=keyboard)
                elif call.message.caption:
                    bot.send_message(call.message.chat.id, constants.rings, reply_markup=keyboard)
            else:  # Остальное расписание - с бд
                try:
                    conn = functions.start_sql()
                    cursor = conn.cursor(buffered=True)
                    functions.isknown(call, pref[2])  # TODO: сделай почеловечести эту функцию

                    get_schedule = "SELECT {day} FROM schedule WHERE klass = '{group}'".format(day=pref[1], group=pref[2])
                    cursor.execute(get_schedule)
                    schedule = cursor.fetchone()[0]
                    cursor.close()
                    conn.close()
                    first_text = "Расписание уроков для группы {group} на {day}".format(group=pref[2], day=dic[pref[1]])
                    keyboard = functions.schedule_inline_keyboard(pref[1], pref[2])
                    bot.send_photo(chat_id=call.message.chat.id, photo=schedule, caption=first_text, reply_markup=keyboard, disable_notification=True)
                except telebot.apihelper.ApiException:
                    traceback.print_exc()
                    ei = "".join(traceback.format_exception(*sys.exc_info()))
                    log('get schedule', ei)
                except:
                    bot.send_message(call.message.chat.id, "Напишите мне в лс (@school134_bot) команду /start")
        # =========================================================================================================== SET Faculty
        elif pref[0] == "g":  # Установка группы
            try:
                conn = functions.start_sql()
                cursor = conn.cursor(buffered=True)
                query = "SELECT klass FROM schedule WHERE klass LIKE '{0}%'".format(pref[1])
                cursor.execute(query)
                klass_sql = cursor.fetchall()
                keyboard_klasses = telebot.types.InlineKeyboardMarkup(row_width=2)
                for i in range(len(klass_sql)):  # Вообще черная магия.
                    klass_sql[i] = str(klass_sql[i]).split("'")[1]
                buttons = [telebot.types.InlineKeyboardButton(text=klass, callback_data='k_' + klass) for klass in klass_sql]
                keyboard_klasses.add(*buttons)
                cursor.close()
                conn.close()
                bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id, text="Из какой вы группы?", reply_markup=keyboard_klasses)
            except:
                traceback.print_exc()
                ei = "".join(traceback.format_exception(*sys.exc_info()))
                log("group_set", ei)
        # =========================================================================================================== SET Group
        elif pref[0] == "k":  # Установка класса
            try:
                conn = functions.start_sql()
                cursor = conn.cursor(buffered=True)
                query = "UPDATE known_users SET klass = '{0}' WHERE id = {1}".format(pref[1], call.from_user.id)
                cursor.execute(query)
                conn.commit()
                cursor.close()
                conn.close()
                bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                      text="Изменения сохранены")
            except:
                traceback.print_exc()
                ei = "".join(traceback.format_exception(*sys.exc_info()))
                log("klass_set", ei)


def set_background(message, data):
    try:
        if message.content_type == "photo":
            photo = message.photo[-1]
            if abs(photo.width - 1280) <= 10 and abs(photo.height - 800) <= 10:
                conn = functions.start_sql()
                cursor = conn.cursor()
                cursor.execute("UPDATE switch SET value = '{0}' WHERE var = 'background'".format(photo.file_id))
                conn.commit()
            else:
                bot.reply_to(message, "Допустимые размеры фото 1280*800 +- 10 пикселей с каждой стороны\n Размер фото:" + str(photo.width) + " " + str(photo.height))
        else:
            bot.reply_to(message, "Отправьте фотографию")
    except:
        traceback.print_exc()
        ei = "".join(traceback.format_exception(*sys.exc_info()))
        name = message.from_user.first_name + ' ' + message.from_user.last_name + ' ' + message.from_user.username + ' ' + str(message.from_user.id) + '\n'
        log('set_background ', name + ei)


def rasp_add(message, data, new_schedule=True):
    try:
        conn = functions.start_sql()
        cursor = conn.cursor()
        path_to_data = 'pic/'
        count_groups = 20
        cur_row = 11

        if new_schedule is True:
            with open(path_to_data + 'rasp.xlsx', 'wb') as f:
                f.write(bot.download_file(bot.get_file(message.document.file_id).file_path))

        wb = openpyxl.load_workbook(path_to_data + 'rasp.xlsx')
        ws = wb.get_active_sheet()
        cur_group = 0
        cur_day = 0

        days = [
            ['mon', 'Понедельник'],
            ['tue', 'Вторник'],
            ['wed', 'Среда'],
            ['thu', 'Четверг'],
            ['fri', 'Пятница'],
            ['sat', 'Суббота']
        ]
        schedule = [
            '8:05-8:45  ',
            '8:55-9:35  ',
            '9:55-10:35 ',
            '10:40-11:20',
            '11:40-12:20',
            '12:30-13:10',
            '13:15-13:55',
            '14:00-14:40',
            '14:45-15:25'
        ]

        lessons = [["", "", ""] for x in range(count_groups)]

        r = re.compile(r'\(.+\)')
        for i in range(5, 44, 2):
            gr = ws.cell(row=10, column=i).value
            po = r.search(gr)
            lessons[cur_group][0] = gr[po.start() + 1: po.end() - 1]
            cur_group += 1
            # заполнил массив групп

        base = Image.open(path_to_data + "back.jpg").convert('RGBA')
        font_group = ImageFont.truetype(path_to_data + "Phenomena-ExtraBold.ttf", 80)
        font_lessons = ImageFont.truetype(path_to_data + "Phenomena-Regular.ttf", 40)
        font_day = ImageFont.truetype(path_to_data + "Phenomena-Regular.ttf", 40)

        while True:
            cur_group = 0
            for i in range(5, 44, 2):
                subject = ws.cell(row=cur_row, column=i).value
                if subject is not None:
                    temp = subject
                    subject = re.sub(r'\s*[2]\s*гр\s*', '', subject)
                    new = ""
                    if subject != temp:  # Если найден шаблон "2 группа", то заменить "1 группа" на "/" Иначе - на ""
                        new = "/"
                    subject = re.sub(r'\s*[1]\s*гр\s*', new, subject)
                    lessons[cur_group][1] += subject + "\n"
                    classroom = ws.cell(row=cur_row, column=i + 1).value
                    if classroom is not None:
                        classroom = re.sub(r"\s+", '/', str(classroom), count=1)
                        lessons[cur_group][2] += classroom
                    lessons[cur_group][2] += " \n"
                cur_group += 1
            cur_row += 1

            cur_lesson = ws.cell(row=cur_row, column=2).value
            next_lsn = ws.cell(row=cur_row + 1, column=2).value

            if next_lsn is None or cur_lesson == 1:

                for i in lessons:
                    group = i[0]
                    lsns = list(zip(i[1].split('\n'), i[2].split('\n')))
                    day = days[cur_day]
                    txt = Image.new("RGBA", base.size, (255, 255, 255, 0))
                    d = ImageDraw.Draw(txt)
                    d.text((10, 745), day[1], (255, 255, 255), font_day)
                    d.text((150, 40), group, (255, 255, 255), font_group)
                    y = 150
                    lesson_number = 1
                    for j in lsns:
                        if j[0] != '':
                            d.text((450, y), str(lesson_number) + '. ' + j[0], (255, 255, 255), font_lessons)
                            d.text((1100, y), j[1], (255, 255, 255), font_lessons)
                            d.text((150, y), schedule[lesson_number - 1], (255, 255, 255), font_lessons)
                            lesson_number += 1
                            y += 70
                        else:
                            break
                    out = Image.alpha_composite(base, txt)
                    # out.save('{0}/{1}.png'.format(day[0], group), 'PNG')
                    out.save(path_to_data + 'temp.png', 'PNG')
                    time.sleep(2)
                    with open(path_to_data + 'temp.png', 'rb') as f:
                        photo = bot.send_photo(-1001079895757, f, disable_notification=True).photo[-1].file_id
                        cursor.execute("UPDATE schedule SET {day} = '{photo_id}' WHERE klass = '{group}'".format(day=day[0],
                                                                                                                 photo_id=photo,
                                                                                                                 group=group.replace('-', '')))
                        # klass строго должен подходить под regexp r'[А-Я]{1,2}\d+'
                        # создать фотку и кинуть в бота
                bot.send_message(message.chat.id, "Расписание на " + days[cur_day][1] + " создано", disable_notification=True)
                cur_day += 1
                for i in lessons:
                    i[1] = ''  # обнулил уроки и classrooms
                    i[2] = ''
            if next_lsn is None:
                break
        conn.commit()
        cursor.close()
        conn.close()
        bot.reply_to(message, "Расписание изменено")
    except:
        traceback.print_exc()
        ei = "".join(traceback.format_exception(*sys.exc_info()))
        name = message.from_user.first_name + ' ' + message.from_user.last_name + ' ' + message.from_user.username + ' ' + str(message.from_user.id) + '\n'
        log('switch ', name + ei)


def ras_switch(message, data):
    try:
        data = data.split('_')  # ras + sql Where expression
        conn = functions.start_sql()
        cursor = conn.cursor(buffered=True)
        try:
            cursor.execute("SELECT id FROM known_users WHERE {query}".format(query=data[1]))
        except Exception as err:
            log('ras_switch', "Initiator: {f_name} {l_name} {u_name}\nInvalid SQL query:\n"
                              "SELECT id FROM known_users WHERE {query}\n{err}".format(f_name=message.from_user.first_name,
                                                                                       l_name=message.from_user.last_name,
                                                                                       u_name=message.from_user.username,
                                                                                       err=str(err),
                                                                                       query=data[1]))
        for record in functions.iter_row(cursor, 100):
            u_id = record[0]
            try:
                if message.content_type == 'text':
                    if message.text[:2] == "nm":
                        text = message.text.replace('nm ', '')
                        bot.send_message(u_id, text)
                    else:
                        bot.send_message(u_id, message.text, parse_mode='Markdown')
                if message.content_type == 'photo':
                    bot.send_photo(u_id, message.photo[-1].file_id, caption=message.caption)
                if message.content_type == 'document':
                    bot.send_document(u_id, message.document.file_id)
                if message.content_type == 'sticker':
                    bot.send_sticker(u_id, message.sticker.file_id)
                if message.content_type == 'video':
                    bot.send_video(u_id, message.video.file_id, caption=message.caption)
                if message.content_type == 'voice':
                    bot.send_voice(u_id, message.voice.file_id)
                if message.content_type == 'contact':
                    bot.send_contact(u_id, message.contact.phone_number, message.contact.first_name, message.contact.last_name)
                if message.content_type == 'location':
                    bot.send_location(u_id, latitude=message.location.latitude, longitude=message.location.longitude)
                if message.content_type == 'audio':
                    bot.send_audio(u_id, message.audio.file_id)
                bot.send_message(constants.main_adm, 'message sent to ' + str(u_id), disable_notification=True)
            except telebot.apihelper.ApiException as APIerr:
                if APIerr.result.status_code == 403:
                    cursor.execute("DELETE FROM `known_users` WHERE id = {uid}".format(uid=u_id))
                    conn.commit()
                    bot.send_message(message.chat.id, '<i>Пользователь\n' + str(u_id) + ' Удален из базы</i>', parse_mode="HTML")
                else:
                    temp = bot.send_message(message.chat.id, '<b>Пользователь\n' + str(u_id) + '\n</b>Err:\n' + str(APIerr), parse_mode='HTML')
            finally:
                time.sleep(0.04)
        bot.send_message(message.from_user.id, "Все сообщения отправлены")
    except:
        traceback.print_exc()
        ei = "".join(traceback.format_exception(*sys.exc_info()))
        name = message.from_user.first_name + ' ' + message.from_user.last_name + ' ' + message.from_user.username + ' ' + str(message.from_user.id) + '\n'
        log('ras ', name + ei)

if __name__ == '__main__':
    application.run()

'''
if __name__ == '__main__':
    try:
        bot.polling(none_stop=True, timeout=5)
    except Exception as e:
        traceback.print_exc()
'''

'''
# месяц возвращается в виде next past this (alphabet order)
server = Flask(__name__)


# server.config['PROPAGATE_EXCEPTIONS'] = True


@server.route("/webhook", methods=['POST'])
def getMessage():
    try:
        update = telebot.types.Update.de_json(request.stream.read().decode("utf-8"))
        bot.process_new_updates([update])
        return "!", 200
    except:
        traceback.print_exc()
        ei = "".join(traceback.format_exception(*sys.exc_info()))
        log('getting msg by hook', ei)


@server.route('/setwebhook')
def webhook():
    try:
        bot.remove_webhook()
        time.sleep(2)
        bot.set_webhook(url="https://python-schoolbot.7e14.starter-us-west-2.openshiftapps.com/webhook")
        return "!", 200
    except:
        traceback.print_exc()
        ei = "".join(traceback.format_exception(*sys.exc_info()))
        log('setwebhook', ei)


if __name__ == '__main__':
    server.run()

'''
